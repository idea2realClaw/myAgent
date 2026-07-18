#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
涨停板复盘报告生成脚本

生成Markdown格式的复盘报告，包含：
1. 今日涨停概况
2. 板块分布
3. 客观评分TOP20
4. 操作建议
"""

import os
from datetime import datetime
from typing import List, Dict


def format_deep_analysis_section(deep_results: List[Dict]) -> List[str]:
    """格式化深度分析章节"""
    lines = []

    if not deep_results:
        return lines

    lines.extend([
        f"## 五、TOP10深度分析（基本面 + 概念 + 涨停原因）",
        f"",
        f"> 对评分最高的10只股票进行基本面、概念板块、资金流向综合分析，给出操作建议。",
        f"",
    ])

    for i, r in enumerate(deep_results, 1):
        fin = r.get('financial_data', {})
        fin_a = r.get('fin_analysis', {})
        fund = r.get('fund_flow', {})

        # 基本面摘要行
        fin_items = []
        if fin.get('pe_ttm') is not None:
            fin_items.append(f"PE {fin['pe_ttm']:.1f}")
        if fin.get('pb') is not None:
            fin_items.append(f"PB {fin['pb']:.2f}")
        if fin.get('roe') is not None:
            fin_items.append(f"ROE {fin['roe']:.1f}%")
        if fin.get('revenue_yoy') is not None:
            fin_items.append(f"营收{fin['revenue_yoy']:+.1f}%")
        if fin.get('profit_yoy') is not None:
            fin_items.append(f"净利{fin['profit_yoy']:+.1f}%")
        fin_summary = ' | '.join(fin_items) if fin_items else '数据缺失'

        # 概念
        concepts = r.get('concepts', [])
        concept_str = ', '.join(concepts[:5]) if concepts else '（未获取到概念数据）'

        # 资金（单位：万元）
        main_in = fund.get('main_inflow', 0)
        if abs(main_in) >= 10000:
            fund_str = f"主力{main_in/10000:+.1f}亿"
        elif abs(main_in) >= 100:
            fund_str = f"主力{main_in/10000:+.2f}亿"
        else:
            fund_str = f"主力{main_in:+.0f}万"

        # 基本面评分条
        fin_total = fin_a.get('total', 0)
        bar_len = 20
        filled = int(fin_total / 100 * bar_len)
        bar = '█' * filled + '░' * (bar_len - filled)

        lines.extend([
            f"### {i}. {r['name']}（{r['code']}） 评分{r['score']}分 {r['grade']}级",
            f"",
            f"**行业:** {r.get('industry', '未知')}  |  **概念:** {concept_str}",
            f"**基本面:** {fin_summary}  |  **资金:** {fund_str}",
            f"**基本面评分:** [{bar}] {fin_total}/100",
            f"",
        ])

        # 基本面详情
        if fin_a.get('comments'):
            lines.append("**基本面评价:**")
            for c in fin_a['comments']:
                lines.append(f"- {c}")
            lines.append("")

        # 涨停原因
        reason = r.get('reason', '')
        if reason:
            lines.extend([
                f"**涨停原因:**",
                reason,
                f"",
            ])

        # 操作建议
        advice = r.get('advice', '')
        if advice:
            lines.extend([
                f"**操作建议:**",
                advice,
                f"",
            ])

        lines.append("---")
        lines.append("")

    return lines


def generate_markdown_report(analysis: Dict, output_path: str = None, deep_results: List[Dict] = None) -> str:
    """
    生成Markdown复盘报告

    Args:
        analysis: 分析结果
        output_path: 输出文件路径
        deep_results: TOP10深度分析结果

    Returns:
        报告内容
    """
    today = datetime.now().strftime('%Y-%m-%d')
    today_cn = datetime.now().strftime('%Y年%m月%d日')

    lines = [
        f"# A股涨停板复盘报告",
        f"",
        f"**{today_cn}**",
        f"",
        f"---",
        f"",
    ]

    # 1. 今日概况
    lines.extend([
        f"## 一、今日涨停概况",
        f"",
        f"| 指标 | 数值 |",
        f"|------|------|",
        f"| 第一板涨停总数 | {analysis['total_count']} 只 |",
        f"| 20cm涨停（科创板/创业板）| {analysis['change_20cm_count']} 只 |",
        f"| 10cm涨停（主板）| {analysis['change_10cm_count']} 只 |",
        f"| 科创板/创业板 | {analysis.get('chip_count', 0)} 只 |",
        f"| 主板 | {analysis.get('normal_count', 0)} 只 |",
        f"| 主板平均评分 | {analysis['avg_score']:.1f} 分 |",
        f"",
        f"### 等级分布（主板）",
        f"",
        f"| 等级 | 数量 | 说明 |",
        f"|------|------|------|",
        f"| S级 | {len(analysis['s_class'])} 只 | 极强推荐 |",
        f"| A级 | {len(analysis['a_class'])} 只 | 强推荐 |",
        f"| B级 | {len(analysis['b_class'])} 只 | 中等关注 |",
        f"| C级 | {len(analysis['c_class'])} 只 | 较弱参考 |",
        f"| D级 | {len(analysis['d_class'])} 只 | 建议观望 |",
        f"",
    ])

    # 2. S/A级推荐
    if analysis['s_class'] or analysis['a_class']:
        lines.extend([
            f"## 二、强势股票（S/A级）",
            f"",
        ])

        for s in analysis['s_class'][:10]:
            lines.extend(format_stock_detail(s))

        for s in analysis['a_class'][:10]:
            lines.extend(format_stock_detail(s))

    # 3. 评分TOP20
    lines.extend([
        f"## 三、客观评分TOP20",
        f"",
        f"**评分标准（100分制，三维度）：**",
        f"",
        f"| 维度 | 满分 | 说明 |",
        f"|------|------|------|",
        f"| 封板强度 | 40分 | 封单金额/流通市值，≥15%满分 |",
        f"| 换手健康 | 30分 | 主板3-8%满分，科创/创业板5-15%满分 |",
        f"| 行业赛道 | 30分 | S级主线30分，A级热点25分，B级20分，C级10分 |",
        f"",
        f"| 排名 | 代码 | 名称 | 封板 | 换手 | 赛道 | 总分 | 等级 | 仓位 |",
        f"|------|------|------|------|------|------|------|------|------|",
    ])

    for s in analysis['all_stocks'][:20]:
        position = s.get('仓位建议', '-')
        chip_tag = " 🔬" if s.get('is_chip') else ""
        lines.append(
            f"| {s['排名']} | {s['code']} | {s['name']}{chip_tag} | "
            f"{s['封板强度']}/40 | {s['换手健康']}/30 | {s['行业赛道']}/30 | "
            f"**{s['总分']}/100** | {s['等级']} | {position} |"
        )

    # 3.5 科创板/创业板板块
    chip_stocks = analysis.get('chip_stocks', [])
    if chip_stocks:
        lines.extend([
            f"",
            f"## 三-B、科创板/创业板涨停股",
            f"",
            f"> 科创板/创业板波动较大（20cm），风险高于主板，仓位需严格控制。",
            f"",
            f"| 排名 | 代码 | 名称 | 封板 | 换手 | 赛道 | 总分 | 等级 | 仓位 |",
            f"|------|------|------|------|------|------|------|------|------|",
        ])
        for s in chip_stocks:
            position = s.get('仓位建议', '-')
            lines.append(
                f"| {s['排名']} | {s['code']} | {s['name']} | "
                f"{s['封板强度']}/40 | {s['换手健康']}/30 | {s['行业赛道']}/30 | "
                f"**{s['总分']}/100** | {s['等级']} | {position} |"
            )

    # 4. 操作建议
    lines.extend([
        f"",
        f"## 四、操作建议",
        f"",
    ])

    # 按仓位分类
    high_positions = [s for s in analysis['all_stocks'] if s.get('仓位比例', 0) >= 0.25]
    mid_positions = [s for s in analysis['all_stocks'] if 0.1 <= s.get('仓位比例', 0) < 0.25]
    low_positions = [s for s in analysis['all_stocks'] if 0 < s.get('仓位比例', 0) < 0.1]

    if high_positions:
        lines.extend([
            f"### 重仓/半仓建议",
            f"",
        ])
        for s in high_positions[:5]:
            lines.append(f"- **{s['name']}** ({s['code']}): {s['总分']}分, {s['仓位建议']} {s['仓位比例']*100:.0f}%")

    if mid_positions:
        lines.extend([
            f"",
            f"### 轻仓关注",
            f"",
        ])
        for s in mid_positions[:5]:
            lines.append(f"- {s['name']} ({s['code']}): {s['总分']}分, {s['仓位建议']} {s['仓位比例']*100:.0f}%")

    if low_positions:
        lines.extend([
            f"",
            f"### 极轻仓/观望",
            f"",
        ])
        for s in low_positions[:5]:
            lines.append(f"- {s['name']} ({s['code']}): {s['总分']}分, {s['仓位建议']}")

    # 5. TOP5深度分析
    if deep_results:
        lines.extend(format_deep_analysis_section(deep_results))

    # 6. 风险提示
    lines.extend([
        f"",
        f"---",
        f"",
        f"## 风险提示",
        f"",
        f"1. 本报告仅供参考，不构成投资建议",
        f"2. 涨停板风险较高，追板需谨慎",
        f"3. 凯利公式仅作为仓位参考，实际仓位应根据个人风险承受能力调整",
        f"4. 市场有风险，投资需谨慎",
        f"",
        f"---",
        f"",
        f"*报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*",
    ])

    report = '\n'.join(lines)

    # 保存文件
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"报告已保存至: {output_path}")

    return report


def format_stock_detail(stock: Dict) -> List[str]:
    """格式化单个股票详情"""
    sealed = stock.get('sealed_amount', 0) / 1e8
    cap = stock.get('float_cap_yi', 0)
    kelly = stock.get('kelly_result', {})

    lines = [
        f"### {stock['name']} ({stock['code']}) [{stock['等级']}级]",
        f"",
        f"| 项目 | 数据 |",
        f"|------|------|",
        f"| 涨跌幅 | {stock['change_pct']:+.2f}% |",
        f"| 换手率 | {stock['turnover']:.2f}% |",
        f"| 封板强度 | {stock['封板强度']}/40 ({stock['封板评价']}) |",
        f"| 换手健康 | {stock['换手健康']}/30 ({stock['换手评价']}) |",
        f"| 行业赛道 | {stock['行业赛道']}/30 ({stock['赛道评价']}) |",
        f"| 综合评分 | **{stock['总分']:.0f}/100分** |",
        f"|",
        f"| 仓位建议 | {stock.get('仓位建议', '-')} ({stock.get('仓位比例', 0)*100:.0f}%) |",
        f"| 风险等级 | {stock.get('风险等级', '-')} |",
        f"",
    ]

    return lines


def generate_excel_report(analysis: Dict, output_path: str = None):
    """
    生成Excel报告（可选）

    需要openpyxl库
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("需要安装openpyxl库来生成Excel报告")
        print("pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "涨停板复盘"

    # 标题
    ws['A1'] = "A股涨停板复盘报告"
    ws['A1'].font = Font(size=16, bold=True)

    # 表头
    headers = ['排名', '代码', '名称', '涨跌幅', '换手率', '封单(亿)', '流通市值(亿)',
               '评分', '等级', '仓位建议']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # 数据
    for row, stock in enumerate(analysis['all_stocks'][:50], 4):
        ws.cell(row=row, column=1, value=stock['排名'])
        ws.cell(row=row, column=2, value=stock['code'])
        ws.cell(row=row, column=3, value=stock['name'])
        ws.cell(row=row, column=4, value=stock['change_pct'])
        ws.cell(row=row, column=5, value=stock['turnover'])
        ws.cell(row=row, column=6, value=stock.get('sealed_amount', 0) / 1e8)
        ws.cell(row=row, column=7, value=stock.get('float_cap_yi', 0))
        ws.cell(row=row, column=8, value=stock['总分'])
        ws.cell(row=row, column=9, value=stock['等级'])
        ws.cell(row=row, column=10, value=stock.get('仓位建议', '-'))

    # 调整列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 12

    # 保存
    if output_path:
        wb.save(output_path)
        print(f"Excel报告已保存至: {output_path}")


if __name__ == '__main__':
    # 测试
    test_analysis = {
        'total_count': 50,
        'change_20cm_count': 20,
        'change_10cm_count': 30,
        'avg_score': 58.5,
        's_class': [],
        'a_class': [],
        'b_class': [],
        'c_class': [],
        'd_class': [],
        'all_stocks': [
            {
                '排名': 1,
                'code': '002361',
                'name': '神剑股份',
                'change_pct': 10.0,
                'turnover': 33.14,
                'sealed_amount': 836000000,
                'float_cap_yi': 110.36,
                '总分': 85,
                '等级': 'S',
                '封单评价': '很强',
                '换手评价': '异常',
                '市值评价': '偏大',
                '仓位建议': '轻仓',
                '仓位比例': 0.15,
                '风险等级': '中等风险',
            }
        ]
    }

    report = generate_markdown_report(test_analysis)
    print(report)
