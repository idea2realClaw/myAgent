#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
最高分涨停股次日跟踪模块

功能：
1. 从数据库读取前一交易日评分最高的涨停股
2. 通过东方财富API获取该股今日开盘价和收盘价
3. 计算开盘买入 -> 收盘的涨幅
4. 记录到xlsx表格（追加模式）
"""

import json
import os
import sys
from datetime import datetime, timedelta
from typing import Optional, Dict, List

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from limit_up_db import LimitUpDB


# xlsx 输出路径
DEFAULT_XLSX_PATH = os.path.expanduser('~/ClawData/WorkBuddy/top-stock-tracking.xlsx')


def get_previous_trading_day(db: LimitUpDB, today_str: str) -> Optional[str]:
    """
    获取前一交易日（在数据库中有数据且不是今天的最近一天）
    
    Args:
        db: 数据库实例
        today_str: 今天日期 YYYY-MM-DD
    
    Returns:
        前一交易日日期字符串，或 None
    """
    all_days = db.get_all_days()
    for d in all_days:
        if d < today_str:
            return d
    return None


def get_top_stock_of_day(db: LimitUpDB, date_str: str) -> Optional[Dict]:
    """
    获取某日评分最高的涨停股
    
    Args:
        db: 数据库实例
        date_str: 日期 YYYY-MM-DD
    
    Returns:
        最高分股票的完整信息，或 None
    """
    day_data = db.get_day(date_str)
    if not day_data:
        return None
    
    top_score = day_data.get('top_stock')
    if top_score:
        return top_score
    
    # 如果没有保存 top_stock 字段，从 stocks 中取第一个（旧数据兼容）
    stocks = day_data.get('stocks', [])
    if stocks:
        # 旧数据：按 change_pct 降序取第一只（粗略）
        stocks_sorted = sorted(stocks, key=lambda s: s.get('change_pct', 0), reverse=True)
        best = stocks_sorted[0]
        return {
            'code': best.get('code', ''),
            'name': best.get('name', ''),
            'change_pct': best.get('change_pct', 0),
            'score': 0,  # 旧数据无评分
            'grade': '?',
        }
    
    return None


def fetch_today_realtime(code: str) -> Optional[Dict]:
    """
    获取某只股票今日行情数据（开盘价、收盘价等）
    优先使用东方财富API，失败后降级到新浪接口（24小时可用）
    
    Returns:
        {'open': 开盘价, 'close': 收盘价, 'high': 最高, 'low': 最低, 'prev_close': 昨收, 'change_pct': 涨跌幅}
    """
    import urllib.request
    
    # 先尝试东方财富API
    result = _fetch_eastmoney(code)
    if result:
        return result
    
    # 降级到新浪接口（24小时可用，含历史收盘数据）
    result = _fetch_sina(code)
    if result:
        return result
    
    return None


def _fetch_eastmoney(code: str) -> Optional[Dict]:
    """东方财富API获取实时数据"""
    import urllib.request
    
    if code.startswith('6'):
        secid = f'1.{code}'
    else:
        secid = f'0.{code}'
    
    url = f'https://push2.eastmoney.com/api/qt/stock/get?secid={secid}&fields=f17,f18,f15,f16,f2,f3,f4&ut=b2884a393a59ad64002292a3e90d46a5'
    
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        response = urllib.request.urlopen(req, timeout=15)
        raw = response.read().decode('utf-8')
        
        data = json.loads(raw).get('data', {})
        if not data:
            return None
        
        return {
            'open': data.get('f17', 0),
            'prev_close': data.get('f18', 0),
            'high': data.get('f15', 0),
            'low': data.get('f16', 0),
            'close': data.get('f2', 0),
            'change_pct': data.get('f3', 0),
        }
    except Exception as e:
        print(f"  [东方财富] 获取 {code} 失败: {e}")
        return None


def _fetch_sina(code: str) -> Optional[Dict]:
    """
    新浪财经API获取行情数据（24小时可用，收盘后也能获取当日数据）
    
    新浪数据格式（逗号分隔）:
    0:名称, 1:今开, 2:昨收, 3:当前价/收盘价, 4:最高, 5:最低, ...
    """
    import urllib.request
    
    if code.startswith('6'):
        prefix = 'sh'
    else:
        prefix = 'sz'
    
    url = f'https://hq.sinajs.cn/list={prefix}{code}'
    
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0',
            'Referer': 'https://finance.sina.com.cn'
        })
        response = urllib.request.urlopen(req, timeout=15)
        raw = response.read().decode('gbk')
        
        # 解析格式: var hq_str_sh605303="园林股份,16.170,16.150,17.770,..."
        if '="' not in raw:
            return None
        
        content = raw.split('="')[1].rstrip('";').rstrip('"')
        fields = content.split(',')
        
        if len(fields) < 10:
            return None
        
        name = fields[0]
        today_open = float(fields[1]) if fields[1] else 0
        prev_close = float(fields[2]) if fields[2] else 0
        close_price = float(fields[3]) if fields[3] else 0
        high = float(fields[4]) if fields[4] else 0
        low = float(fields[5]) if fields[5] else 0
        
        if prev_close > 0:
            change_pct = (close_price - prev_close) / prev_close * 100
        else:
            change_pct = 0
        
        return {
            'open': today_open,
            'prev_close': prev_close,
            'high': high,
            'low': low,
            'close': close_price,
            'change_pct': change_pct,
        }
    except Exception as e:
        print(f"  [新浪] 获取 {code} 失败: {e}")
        return None


def track_top_stock(prev_date: str, today_str: str, db: LimitUpDB) -> Optional[Dict]:
    """
    跟踪前一日最高分股票今日表现
    
    Args:
        prev_date: 前一日日期
        today_str: 今日日期
        db: 数据库实例
    
    Returns:
        跟踪记录字典，或 None（如果无法跟踪）
    """
    top_stock = get_top_stock_of_day(db, prev_date)
    if not top_stock:
        print(f"  [跟踪] 无 {prev_date} 的最高分股票数据")
        return None
    
    code = top_stock.get('code', '')
    name = top_stock.get('name', '')
    prev_change_pct = top_stock.get('change_pct', 0)
    score = top_stock.get('score', 0)
    grade = top_stock.get('grade', '?')
    
    print(f"  [跟踪] 前日最高分: {name}({code}) 评分{score}分 {grade}级 涨停{prev_change_pct:+.1f}%")
    
    # 获取今日实时数据
    realtime = fetch_today_realtime(code)
    if not realtime:
        print(f"  [跟踪] 无法获取今日实时数据，跳过")
        return None
    
    today_open = realtime.get('open', 0)
    today_close = realtime.get('close', 0)
    today_prev_close = realtime.get('prev_close', 0)
    today_change_pct = realtime.get('change_pct', 0)
    
    # 计算开盘买入到收盘的涨幅
    # 即：今日开盘价买入，收盘时盈亏 = (收盘价 - 开盘价) / 开盘价 * 100
    if today_open > 0 and today_close > 0:
        open_to_close_pct = (today_close - today_open) / today_open * 100
    else:
        open_to_close_pct = 0
    
    # 计算相对前日涨停收盘价的涨幅（隔夜持仓收益）
    # 前日涨停收盘价 ≈ 前日 prev_close * (1 + change_pct/100)
    # 但更准确的是用今日 prev_close（即前日收盘价）
    if today_prev_close > 0:
        overnight_profit = (today_open - today_prev_close) / today_prev_close * 100
        total_profit = (today_close - today_prev_close) / today_prev_close * 100
    else:
        overnight_profit = 0
        total_profit = 0
    
    record = {
        'prev_date': prev_date,
        'stock_name': name,
        'stock_code': code,
        'prev_score': score,
        'prev_grade': grade,
        'prev_change_pct': prev_change_pct,
        'today_date': today_str,
        'today_open': today_open,
        'today_close': today_close,
        'today_change_pct': today_change_pct,
        'open_to_close_pct': open_to_close_pct,
        'overnight_profit': overnight_profit,
        'total_profit': total_profit,
    }
    
    print(f"  [跟踪] 今日: 开盘{today_open} 收盘{today_close} 涨幅{today_change_pct:+.1f}%")
    print(f"  [跟踪] 开盘->收盘: {open_to_close_pct:+.2f}%")
    print(f"  [跟踪] 隔夜溢价: {overnight_profit:+.2f}%  总收益: {total_profit:+.2f}%")
    
    return record


def save_to_xlsx(records: List[Dict], xlsx_path: str = None):
    """
    将跟踪记录保存/追加到xlsx文件
    
    Args:
        records: 跟踪记录列表
        xlsx_path: xlsx文件路径
    """
    if xlsx_path is None:
        xlsx_path = DEFAULT_XLSX_PATH
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("[xlsx] 需要安装 openpyxl: pip install openpyxl")
        # 降级为CSV
        save_to_csv(records, xlsx_path.replace('.xlsx', '.csv'))
        return
    
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    
    headers = [
        '前一日日期', '股票名称', '股票代码', '前一日评分', '前一日等级',
        '前一日涨幅%', '今日日期', '今日开盘价', '今日收盘价',
        '今日涨幅%', '开盘买入->收盘涨幅%', '隔夜溢价%', '总收益%'
    ]
    keys = [
        'prev_date', 'stock_name', 'stock_code', 'prev_score', 'prev_grade',
        'prev_change_pct', 'today_date', 'today_open', 'today_close',
        'today_change_pct', 'open_to_close_pct', 'overnight_profit', 'total_profit'
    ]
    
    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    if os.path.exists(xlsx_path):
        # 追加模式：读取现有数据，检查是否已有今日记录
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # 获取已有记录的 (prev_date, stock_code) 集合，避免重复
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2]:
                existing.add((str(row[0]), str(row[2])))
        
        new_count = 0
        for rec in records:
            key = (rec['prev_date'], rec['stock_code'])
            if key not in existing:
                ws.append([rec.get(k, '') for k in keys])
                existing.add(key)
                new_count += 1
        
        # 重新应用样式（对新追加的行）
        for row_idx in range(ws.max_row - new_count + 1, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align
                
                # 涨幅列着色
                if col_idx in [6, 10, 11, 12, 13]:  # 涨幅相关列
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = positive_fill
                        elif cell.value < 0:
                            cell.fill = negative_fill
        
        wb.save(xlsx_path)
        print(f"[xlsx] 追加 {new_count} 条记录到 {xlsx_path}")
        
    else:
        # 新建文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "最高分股票跟踪"
        
        # 写表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # 写数据
        for rec in records:
            row_values = [rec.get(k, '') for k in keys]
            ws.append(row_values)
        
        # 应用样式
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align
                
                if col_idx in [6, 10, 11, 12, 13]:
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = positive_fill
                        elif cell.value < 0:
                            cell.fill = negative_fill
        
        # 调整列宽
        col_widths = [12, 12, 10, 10, 8, 10, 12, 12, 12, 10, 16, 10, 10]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        wb.save(xlsx_path)
        print(f"[xlsx] 新建文件并写入 {len(records)} 条记录到 {xlsx_path}")


def save_to_csv(records: List[Dict], csv_path: str):
    """降级方案：保存为CSV"""
    import csv
    
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    
    headers = [
        '前一日日期', '股票名称', '股票代码', '前一日评分', '前一日等级',
        '前一日涨幅%', '今日日期', '今日开盘价', '今日收盘价',
        '今日涨幅%', '开盘买入->收盘涨幅%', '隔夜溢价%', '总收益%'
    ]
    keys = [
        'prev_date', 'stock_name', 'stock_code', 'prev_score', 'prev_grade',
        'prev_change_pct', 'today_date', 'today_open', 'today_close',
        'today_change_pct', 'open_to_close_pct', 'overnight_profit', 'total_profit'
    ]
    
    file_exists = os.path.exists(csv_path)
    
    with open(csv_path, 'a', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        for rec in records:
            writer.writerow([rec.get(k, '') for k in keys])
    
    print(f"[csv] 保存 {len(records)} 条记录到 {csv_path}")


def run_tracking(today_str: str = None, db: LimitUpDB = None, xlsx_path: str = None):
    """
    主跟踪函数：获取前一日最高分股票，跟踪今日表现，写入xlsx
    
    Args:
        today_str: 今日日期，默认今天
        db: 数据库实例，默认新建
        xlsx_path: xlsx输出路径，默认 ~/ClawData/WorkBuddy/top-stock-tracking.xlsx
    
    Returns:
        跟踪记录字典，或 None
    """
    if today_str is None:
        today_str = datetime.now().strftime('%Y-%m-%d')
    if db is None:
        db = LimitUpDB()
    
    print(f"\n[跟踪] 开始跟踪前一日最高分股票 -> 今日表现")
    print(f"[跟踪] 今日: {today_str}")
    
    # 1. 找到前一交易日
    prev_date = get_previous_trading_day(db, today_str)
    if not prev_date:
        print(f"[跟踪] 数据库中无前一日数据，跳过")
        return None
    
    print(f"[跟踪] 前一交易日: {prev_date}")
    
    # 2. 获取前一日最高分股票
    record = track_top_stock(prev_date, today_str, db)
    if not record:
        return None
    
    # 3. 写入xlsx
    save_to_xlsx([record], xlsx_path)
    
    return record


if __name__ == '__main__':
    run_tracking()
